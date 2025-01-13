import csv
import os
from re import X
import matplotlib.pyplot as plt
import numpy as np
import statistics
import pandas as pd
import openpyxl
import math

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

x_values = []
y_values = []

x_values = [0, 6.364, 6.44, 6.516, 8.216, 8.5095, 8.803, 9.916, 10.578, 11.24]
y_values = [-245.6, -245.3, -245, -243.6, -243.2, -242.85, -242.13, -241.6, -241.1, -240.7]
offset_values = list(itertools.product(x_values, y_values)) # All combinations of x and y values


def calculate_errors(ax, ay, bx, by):
    ax_error = np.abs(ax - bx) / ax
    ay_error = np.abs(ay - by) / ay
    bx_error = np.abs(bx1 - bx2) / bx2
    by_error = np.abs(by1 - by2) / by2

    ax_error_avg = np.mean(ax_error)
    ay_error_avg = np.mean(ay_error)
    bx_error_avg = np.mean(bx_error)
    by_error_avg = np.mean(by_error)

    return ax_error, ay_error, bx_error, by_error, ax_error_avg, ay_error_avg, bx_error_avg, by_error_avg


def rotate_array(ax1, ay1, rot_matrix, isShape, shapeVector):
    # Update the rotation matrix with the new theta
    # rot_matrix = np.array([[np.cos(theta), -np.sin(theta)], [np.sin(theta), np.cos(theta)]])

    # Stack ax1 and ay1 vertically to create a 2D array of coordinates
    a_coords = np.vstack((ax1, ay1))

    # Apply the rotation matrix to the coordinates
    a_rotated = np.matmul(rot_matrix, a_coords)
    trans_vector = np.array([a_rotated[0,0]*-1, (a_rotated[1][0] * -1) - 50])
    if not isShape:
        trans_vector = shapeVector
    # Apply the translation vector to the rotated coordinates
    a_transformed = a_rotated + np.expand_dims(trans_vector, axis=1)

    # Split the transformed coordinates back into separate arrays
    t_x, t_y = a_transformed[0], a_transformed[1]

    if isShape:
        return t_x, t_y, trans_vector
    else:
        return t_x, t_y



#Very Small Value to avoid dividing by zero during error calculations
epsilon = 0.0001
for iterationNum in range(100):
    for shapeNum in range(3):
        x_pose = []
        y_pose = []
        z_pose = []
        yaw = []
        pitch = []
        roll = []

        static_x_pose = []
        static_y_pose = []
        static_z_pose = []

        # PULL VALUES FROM POSE CSV
        filename1 = offset_values[(iterationNum*3)+shapeNum]+'_square'+str(shapeNum+1)+'_r'+str(iterationNum+1)+'_t2.csv'
        filename2 = offset_values[(iterationNum*3)+shapeNum]+'_square'+str(shapeNum+1)+'_r'+str(iterationNum+1)+'_t1.csv'
        file_split = filename1.split('_')

        with open('output/'+filename1) as csvfile:
            csv_reader = csv.reader(csvfile)
            next(csv_reader)  # Skip the header row if there is one

            for row in csv_reader:
                if row:
                    x_pose.append(float(row[0]))  # Assuming data in the first column are floats
                    y_pose.append(float(row[2]))  # Assuming data in the second column are floats
                    z_pose.append(float(row[1]))  # Assuming data in the first column are floats
                    yaw.append(float(row[4]))  # Assuming data in the second column are floats
                    pitch.append(float(row[3]))  # Assuming data in the second column are floats
                    roll.append(float(row[5]))  # Assuming data in the second column are floats

        with open('output/'+filename2) as csvfile:
            csv_reader = csv.reader(csvfile)
            next(csv_reader)  # Skip the header row if there is one

            for row in csv_reader:
                if row:
                    static_x_pose.append(float(row[0]))  # Assuming data in the first column are floats
                    static_y_pose.append(float(row[2]))  # Assuming data in the second column are floats
                    static_z_pose.append(float(row[1]))  # Assuming data in the first column are floats

        # MAKE SURE TO MULTIPLY X,Y,Z BY 1000 TO GET IN MM
        # WE MAY HAVE TO MULTIPLY X VALUES BY -1

        # Create NumPy arrays
        x_pose = np.array(x_pose)
        y_pose = np.array(y_pose)
        z_pose = np.array(z_pose)
        yaw = np.array(yaw)
        pitch = np.array(pitch)
        roll = np.array(roll)

        x_pose = x_pose * 1000
        y_pose = y_pose * 1000
        z_pose = z_pose * 1000

        static_x_pose = np.array(static_x_pose)
        static_y_pose = np.array(static_y_pose)
        static_z_pose = np.array(static_z_pose)
        static_x_pose = static_x_pose * 1000
        static_y_pose = static_y_pose * 1000
        static_z_pose = static_z_pose * 1000


        # CHANGE TO PULL VALUES FROM GCODE
        # Read the GCODE file
        filename3 = 'test_square.gcode'

        with open(filename3, 'r') as file:
            gcode = file.readlines()

        # Extract X and Y coordinates
        coords = np.array(
            [[float(value[1:]) for value in line.split() if value.startswith('X') or value.startswith('Y')] for line in gcode if
             line.startswith('G0')])

        #print(coords)
        # Transpose the coordinates array
        coords_array = coords.T
        print(coords_array)


        # New attempt at precise calibration theta = -30.44
        # OG Calibration is set to -30.45
        # Define a rotation angle of theta degrees
        #Pitch: 78.8591078645256
        theta = yaw[0]-90-180
        # Convert theta from degrees to radians
        theta = np.radians(theta)
        # -30.45
        # Define a rotation matrix and a translation vector
        rot_matrix = np.array([[np.cos(theta), -np.sin(theta)], [np.sin(theta), np.cos(theta)]])


        # Rotate and translate the array using the function
        t_xpose, t_ypose, shape_vector = rotate_array(x_pose, y_pose, rot_matrix, True, [])
        st_xpose, st_ypose = rotate_array(static_x_pose, static_y_pose, rot_matrix, False, shape_vector)
        min_y = t_ypose.min()
        min_x = t_xpose[0]



        print("Min Y: " + str(min_y))
        print(np.where(t_ypose == t_ypose.min()))
        print("X at Min Y" + str(min_x))

        #TEST TO FIND AVERAGE TRACKED POINTS PER 10MM: 25
        point = 0
        for i in range(t_ypose.size):
            if (t_ypose[i] < -10) and (t_ypose[i] > -30):
                point += 1

        print("Points:" + str(point))

        total_tracked_points = t_ypose.size
        section_points = round(total_tracked_points/4)
        print(section_points)

        #(0, -50) ---> (0, 50)
        coords_array_X = np.linspace(coords_array[0][0], coords_array[0][1], section_points)
        coords_array_Y = np.linspace(coords_array[1][0], coords_array[1][1], section_points)

        #(0, 50) ---> (0,0)
        coords_array_X = np.append(coords_array_X, np.linspace(coords_array[0][1], coords_array[0][2], section_points))
        coords_array_Y = np.append(coords_array_Y, np.linspace(coords_array[1][1], coords_array[1][2], section_points))

        #(0,0) ---> (-50, 0)
        coords_array_X = np.append(coords_array_X, np.linspace(coords_array[0][2], coords_array[0][3], section_points))
        coords_array_Y = np.append(coords_array_Y, np.linspace(coords_array[1][2], coords_array[1][3], section_points))

        #(-50, 0) ---> (50, 0)
        coords_array_X = np.append(coords_array_X, np.linspace(coords_array[0][3], coords_array[0][4], section_points))
        coords_array_Y = np.append(coords_array_Y, np.linspace(coords_array[1][3], coords_array[1][4], section_points))

        # Print the coordinates array
        #print(coords_array_Y)
        print(static_x_pose)
        print(st_xpose)
        """
        fig, axs = plt.subplots(nrows=2, ncols=2, figsize=(10, 8))
        plt.subplots_adjust(hspace=0.3)
        #axs[0][0].scatter(coords_array[0], coords_array[1], c='blue')
        axs[0][0].scatter(coords_array_X, coords_array_Y, c='blue')
        # axs[0].scatter(bx2,by2, c='red')
        axs[0][0].set_title('Original')
        axs[0][0].set_aspect('equal')
        axs[0][0].set_xlabel('mm')
        axs[0][0].set_ylabel('mm')


        axs[0][1].scatter(x_pose, y_pose, c='blue')
        axs[0][1].set_title('Original Tracked')
        axs[0][1].set_aspect('equal')
        axs[0][1].set_xlabel('mm')
        axs[0][1].set_ylabel('mm')

        axs[1][0].scatter(t_xpose, t_ypose, c='blue')
        #axs[1][0].scatter(st_xpose, st_ypose, c='red')
        axs[1][0].set_title('Transformed Tracked')
        axs[1][0].set_aspect('equal')
        axs[1][0].set_xlabel('mm')
        axs[1][0].set_ylabel('mm')

        axs[1][1].scatter(coords_array[0], coords_array[1], c='blue')
        axs[1][1].scatter(coords_array_X, coords_array_Y, c='blue')
        axs[1][1].scatter(t_xpose, t_ypose, c='red')
        axs[1][1].set_title('Composite Plot')
        axs[1][1].set_aspect('equal')
        axs[1][1].set_xlabel('mm')
        axs[1][1].set_ylabel('mm')
        """

        #plt.savefig("Visualizations/"+filename1[:13] + "_VISUAL.png")
        #plt.show()
        #plt.close()

        #SIZED ARRAYS FOR COMPARISON
        if t_xpose.size > coords_array_X.size:
            sizeDiff = t_xpose.size - coords_array_X.size
            coords_array_X = np.pad(coords_array_X, (0, sizeDiff), mode='constant')
            coords_array_Y = np.pad(coords_array_Y, (0, sizeDiff), mode='constant', constant_values=(-50,))

        else:
            sizeDiff = coords_array_X.size - t_xpose.size
            t_xpose = np.pad(t_xpose, (0, sizeDiff), mode='constant')
            t_ypose = np.pad(t_ypose, (0, sizeDiff), mode='constant', constant_values=(-50,))


        Per_Error_X = []
        Per_Error_Y = []
        Abs_Error_X = []
        Abs_Error_Y = []
        for i in range(coords_array_X.size):
            #print(i)
            Per_Error_X.append( (( (coords_array_X[i] + epsilon) - (t_xpose[i] + epsilon) ) / (coords_array_X[i] + epsilon))*100 )
            Per_Error_Y.append( (( (coords_array_Y[i] + epsilon) - (t_ypose[i] + epsilon) ) / (coords_array_Y    [i] + epsilon))*100 )

            Abs_Error_X.append(((coords_array_X[i]) - (t_xpose[i])) )
            Abs_Error_Y.append(((coords_array_Y[i]) - (t_ypose[i])) )

        #MEAN SQUARED ERROR
        X_MSE = np.square(np.subtract(t_xpose,coords_array_X)).mean()
        Y_MSE = np.square(np.subtract(t_ypose,coords_array_Y)).mean()
        #MEAN
        #x_mean = np.mean(t_xpose)
        #y_mean = np.mean(t_ypose)
        x_mean = np.mean(Abs_Error_X)
        y_mean = np.mean(Abs_Error_Y)
        #STANDARD DEVIATION
        x_stdev = statistics.stdev(Abs_Error_X)
        y_stdev = statistics.stdev(Abs_Error_Y)
        print(np.mean(Abs_Error_X))


        #SAVING DATA TO EXCEL FILE
        per_point_data = pd.DataFrame({'ax':coords_array_X, 'ay':coords_array_Y, 'bx': t_xpose, 'by': t_ypose,
                                       'x_per_error': Per_Error_X, 'y_per_error': Per_Error_Y,
                                       'x_abs_error': Abs_Error_X, 'y_abs_error': Abs_Error_Y})
        data_set_calc = pd.DataFrame({'x_mse': [X_MSE], 'y_mse': [Y_MSE], 'x_rmse': [np.sqrt(X_MSE)], 'y_rmse': [np.sqrt(Y_MSE)], 'x_mean': x_mean, 'y_mean': y_mean, 'x_stdev': x_stdev, 'y_stdev': y_stdev})

        static_tracker_data = pd.DataFrame({'static_x': st_xpose, 'static_y': st_ypose})
        print('Error_Analysis/'+file_split[0][:5]+'_'+file_split[1]+'_error')


        existing_file = 'Error_Analysis/'+file_split[0]+'_'+file_split[1][:6]+'_'+file_split[2]+'_error.xlsx'

        # Check if the file already exists
        if os.path.isfile(existing_file):
            # Load the existing workbook
            book = load_workbook(existing_file)
            if 'Sheet' in book.sheetnames:
                del book['Sheet']
        else:
            # Create a new workbook
            book = Workbook()

        # Create a Pandas Excel writer using openpyxl engine and existing workbook
        writer = pd.ExcelWriter(existing_file, engine='openpyxl')
        writer.book = book

        # Save the data from each iteration to separate sheets
        per_point_data.to_excel(writer, sheet_name="square" + str(shapeNum + 1), encoding='utf-8')
        data_set_calc.to_excel(writer, sheet_name="square" + str(shapeNum + 1), startcol=9, index=False, encoding='utf-8')
        static_tracker_data.to_excel(writer, sheet_name="static_square" + str(shapeNum + 1), encoding='utf-8')

        """
        if shapeNum == 0:
            fig, axs = plt.subplots(nrows=2, ncols=2, figsize=(12, 12))
            plt.subplots_adjust(hspace=0.3)

            points = np.linspace(1, coords_array_X.size, coords_array_X.size)
            abs_mean_line_X = np.linspace(x_mean, x_mean, coords_array_X.size)
            abs_mean_line_Y = np.linspace(y_mean, y_mean, coords_array_X.size)

            axs[0][0].scatter(points, coords_array_X, c='blue')
            axs[0][0].scatter(points, t_xpose, c='red')
            axs[0][0].set_title('X Traversal')
            axs[0][0].set_xlabel('Tracked Points')
            axs[0][0].set_ylabel('X Value')

            axs[0][1].scatter(points, coords_array_Y, c='blue')
            axs[0][1].scatter(points, t_ypose, c='red')
            axs[0][1].set_title('Y Traversal')
            axs[0][1].set_xlabel('Tracked Points')
            axs[0][1].set_ylabel('Y Value')

            axs[1][0].scatter(points, Abs_Error_X, c='blue')
            axs[1][0].scatter(points, abs_mean_line_X, c='red')
            axs[1][0].set_title('Absolute Error Across X Traversal')
            axs[1][0].set_xlabel('Tracked Points')
            axs[1][0].set_ylabel('Absolute X Error')

            axs[1][1].scatter(points, Abs_Error_Y, c='blue')
            axs[1][1].scatter(points, abs_mean_line_Y, c='red')
            axs[1][1].set_title('Absolute Error Across Y Traversal')
            axs[1][1].set_xlabel('Tracked Points')
            axs[1][1].set_ylabel('Absolute Y Error')
            # plt.show()
            #plt.savefig("Visualizations/" + filename1[:12] + "_ABSOLUTE_ERROR.png")
            plt.close()
        """

        # Save and close the writer
        #writer.save()
        writer.close()